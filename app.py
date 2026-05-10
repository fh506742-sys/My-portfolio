import os
import re
import json
import csv
import secrets
import tempfile
import sqlite3
import io
import random
import operator
import ast
import sys
import time
import numpy as np
from decimal import Decimal, InvalidOperation, getcontext
from sympy import symbols, Eq, solve, simplify, expand, factor, diff, integrate, limit, sympify, SympifyError, lambdify
from sympy.parsing.sympy_parser import parse_expr, standard_transformations, implicit_multiplication_application
import signal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Flask, render_template, request, jsonify, Response, send_file
import requests
from gtts import gTTS
import PyPDF2
from dotenv import load_dotenv
from ddgs import DDGS
from bs4 import BeautifulSoup

# Load environment variables
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"), override=True)

# Set high precision for decimal calculations
getcontext().prec = 50

MATH_CACHE = {}

# ─────────────────────────────────────────────────────────────────────────────
# App Setup
# ─────────────────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.secret_key     = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))
API_KEY            = os.environ.get("GROQ_API_KEY")
GSHEET_CREDS       = os.environ.get("GOOGLE_CREDENTIALS_FILE", "google_credentials.json")
ALLOWED_EXTENSIONS = {".pdf", ".txt", ".md", ".xlsx", ".xls"}
DB_PATH            = "memory.db"
MODEL              = "llama-3.3-70b-versatile"

# Global contexts
bank_sheet_context   = {"summary": None, "raw_preview": None, "headers": [], "records": []}
scraped_data_context = {"rows": [], "headers": [], "title": "", "url": ""}

# Correction workflow state
correction_state = {
    "pending_fixes":  [],
    "accepted_fixes": [],
    "done": False
}

# ─────────────────────────────────────────────────────────────────────────────
# SQLite Memory
# ─────────────────────────────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS messages (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            role      TEXT NOT NULL,
            content   TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

def load_history(limit=20):
    conn = sqlite3.connect(DB_PATH)
    rows = conn.execute(
        "SELECT role, content FROM messages ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return [{"role": row[0], "content": row[1]} for row in reversed(rows)]

def save_message(role, content):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("INSERT INTO messages (role, content) VALUES (?, ?)", (role, content))
    conn.commit()
    conn.close()

def clear_history():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM messages")
    conn.commit()
    conn.close()

init_db()

# ─────────────────────────────────────────────────────────────────────────────
# Excel Bank Sheet Analyzer
# ─────────────────────────────────────────────────────────────────────────────

def parse_excel(file):
    wb       = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws       = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return None, None, None, "Empty sheet."

    raw_headers = all_rows[0]
    headers     = [str(h).strip() if h is not None else f"Col{i}" for i, h in enumerate(raw_headers)]
    data_rows   = all_rows[1:]
    records     = [dict(zip(headers, row)) for row in data_rows if any(c is not None for c in row)]
    total       = len(records)

    def find_col(kws):
        for h in headers:
            if any(k.lower() in h.lower() for k in kws):
                return h
        return None

    loan_col        = find_col(["loan", "principal", "amount", "loan_amount"])
    installment_col = find_col(["installment", "emi", "monthly", "payment"])
    paid_col        = find_col(["paid", "paid_amount", "total_paid", "received"])
    balance_col     = find_col(["balance", "outstanding", "remaining", "due"])
    status_col      = find_col(["status", "state", "loan_status"])
    defaulter_col   = find_col(["default", "defaulter", "overdue", "delinquent"])
    name_col        = find_col(["name", "customer", "borrower", "client", "person"])

    def safe_float(v):
        try:
            return float(str(v).replace(",", "").replace("$", "").replace("৳", "").strip())
        except:
            return None

    stats = {"total_records": total, "columns": headers}

    for col, keys in [
        (loan_col,        ["total_loan_amount", "average_loan_amount", "max_loan", "min_loan"]),
        (balance_col,     ["total_outstanding_balance", "average_balance"]),
        (paid_col,        ["total_paid"]),
        (installment_col, ["total_installment_due", "average_installment"]),
    ]:
        if col:
            vals = [v for v in [safe_float(r.get(col)) for r in records] if v is not None]
            if vals:
                if "loan" in keys[0]:
                    stats.update({keys[0]: sum(vals), keys[1]: sum(vals) / len(vals),
                                   keys[2]: max(vals), keys[3]: min(vals)})
                elif "balance" in keys[0]:
                    stats.update({keys[0]: sum(vals), keys[1]: sum(vals) / len(vals)})
                elif "paid" in keys[0]:
                    stats[keys[0]] = sum(vals)
                elif "installment" in keys[0]:
                    stats.update({keys[0]: sum(vals), keys[1]: sum(vals) / len(vals)})

    defaulter_count = 0
    if defaulter_col:
        for r in records:
            if any(x in str(r.get(defaulter_col, "")).lower() for x in ["yes", "true", "1", "default", "overdue"]):
                defaulter_count += 1
        stats["defaulter_count"] = defaulter_count

    if status_col:
        sc = {}
        for r in records:
            s = str(r.get(status_col, "Unknown")).strip()
            sc[s] = sc.get(s, 0) + 1
        stats["status_breakdown"] = sc
        if not defaulter_col:
            for k, v in sc.items():
                if any(x in k.lower() for x in ["default", "overdue", "late", "delinquent"]):
                    defaulter_count += v
            stats["defaulter_count"] = defaulter_count

    if name_col and balance_col:
        topped = sorted(
            [(r.get(name_col), safe_float(r.get(balance_col))) for r in records if safe_float(r.get(balance_col))],
            key=lambda x: x[1], reverse=True
        )[:10]
        stats["top_10_by_outstanding_balance"] = [{"name": str(n), "balance": b} for n, b in topped]

    preview = "\n".join(
        [" | ".join(headers)] +
        [" | ".join(str(v) if v is not None else "" for v in row) for row in data_rows[:50]]
    )
    return headers, records, stats, preview


def format_stats_for_ai(stats):
    lines = [f"BANK SHEET STATISTICAL SUMMARY ({stats.get('total_records', 0)} total records)"]
    lines.append(f"Columns: {', '.join(stats.get('columns', []))}\n")
    for k, lbl in [
        ("total_loan_amount",         "Total Loan Amount"),
        ("average_loan_amount",       "Average Loan"),
        ("max_loan",                  "Largest Loan"),
        ("min_loan",                  "Smallest Loan"),
        ("total_outstanding_balance", "Total Outstanding"),
        ("average_balance",           "Avg Balance"),
        ("total_paid",                "Total Paid"),
        ("total_installment_due",     "Total Installment Due"),
        ("average_installment",       "Avg Installment"),
    ]:
        if k in stats:
            lines.append(f"{lbl}: {stats[k]:,.2f}")
    if "defaulter_count" in stats:
        pct = stats["defaulter_count"] / stats["total_records"] * 100 if stats["total_records"] else 0
        lines.append(f"\nDefaulters: {stats['defaulter_count']} ({pct:.1f}%)")
    if "status_breakdown" in stats:
        lines.append("\nStatus Breakdown:")
        for k, v in stats["status_breakdown"].items():
            lines.append(f"  {k}: {v}")
    if "top_10_by_outstanding_balance" in stats:
        lines.append("\nTop 10 by Balance:")
        for item in stats["top_10_by_outstanding_balance"]:
            lines.append(f"  {item['name']}: {item['balance']:,.2f}")
    return "\n".join(lines)

# ─────────────────────────────────────────────────────────────────────────────
# AI-Powered Error Detection
# ─────────────────────────────────────────────────────────────────────────────

def detect_errors(headers, records, stats_text, preview):
    sample_rows = []
    for i, r in enumerate(records[:100]):
        row_str = " | ".join(f"{k}: {v}" for k, v in r.items())
        sample_rows.append(f"Row {i+1}: {row_str}")
    sample_text = "\n".join(sample_rows)

    prompt = (
        "You are a meticulous bank data auditor. Analyze this bank sheet data and find ALL mistakes.\n\n"
        "STATISTICS:\n" + stats_text + "\n\n"
        "DATA SAMPLE (first 100 rows):\n" + sample_text + "\n\n"
        "Find these types of errors:\n"
        "1. CALCULATION ERRORS: e.g. outstanding_balance should equal loan_amount minus total_paid.\n"
        "2. SPELLING ERRORS: obvious typos in names, statuses, or text fields.\n"
        "3. IMPOSSIBLE VALUES: negative loan amounts, balances larger than original loan, etc.\n"
        "4. INCONSISTENCY: a person marked 'paid' but still has a large balance, etc.\n\n"
        'Return ONLY a JSON array. Each item must have: '
        '{"row_idx": <1-based>, "column": "<name>", "old_value": "<wrong>", '
        '"new_value": "<corrected>", "reason": "<explanation>", '
        '"type": "<calculation|spelling|impossible|inconsistency>"}\n\n'
        "If no errors found, return []\n"
        "Return ONLY valid JSON. No markdown, no explanation outside the JSON."
    )

    try:
        res = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
            json={
                "model": MODEL,
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 3000,
                "temperature": 0.7,
                "top_p": 0.9,
                "presence_penalty": 0.1,
                "frequency_penalty": 0.1,
            },
            timeout=40
        )
        raw   = res.json()["choices"][0]["message"]["content"].strip()
        raw   = raw.replace("```json", "").replace("```", "").strip()
        fixes = json.loads(raw)
        if not isinstance(fixes, list):
            return []
        for i, fix in enumerate(fixes):
            fix["id"] = i + 1
        return fixes
    except Exception as e:
        print("ERROR DETECTION ERROR:", e)
        return []

# ─────────────────────────────────────────────────────────────────────────────
# Generate Corrected Excel File
# ─────────────────────────────────────────────────────────────────────────────

def build_corrected_excel(headers, records, accepted_fixes):
    corrected  = [dict(r) for r in records]
    change_log = []

    for fix in accepted_fixes:
        row_idx = fix.get("row_idx", 0) - 1
        col     = fix.get("column", "")
        new_val = fix.get("new_value", "")
        old_val = fix.get("old_value", "")
        reason  = fix.get("reason", "")
        if 0 <= row_idx < len(corrected) and col in corrected[row_idx]:
            corrected[row_idx][col] = new_val
            change_log.append({"Row": row_idx + 1, "Column": col,
                                "Old Value": old_val, "New Value": new_val,
                                "Reason": reason, "Type": fix.get("type", "")})

    wb           = openpyxl.Workbook()
    ws           = wb.active
    ws.title     = "Corrected Data"
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1e3a6e")
    fixed_fill   = PatternFill("solid", fgColor="d1fae5")
    fixed_font   = Font(color="065f46", bold=True)
    center_align = Alignment(horizontal="center")

    for col_idx, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    fixed_cells = {(f["row_idx"] - 1, f["column"]) for f in accepted_fixes}

    for row_idx, record in enumerate(corrected):
        for col_idx, h in enumerate(headers, 1):
            val  = record.get(h, "")
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=str(val) if val is not None else "")
            if (row_idx, h) in fixed_cells:
                cell.fill = fixed_fill
                cell.font = fixed_font

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    ws2         = wb.create_sheet("Change Log")
    log_headers = ["Row", "Column", "Old Value", "New Value", "Reason", "Type"]

    for col_idx, h in enumerate(log_headers, 1):
        cell           = ws2.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    for row_idx, entry in enumerate(change_log, 2):
        for col_idx, h in enumerate(log_headers, 1):
            ws2.cell(row=row_idx, column=col_idx, value=str(entry.get(h, "")))

    for col in ws2.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────────────────────────────
# Web Scraper
# ─────────────────────────────────────────────────────────────────────────────

def scrape_url(url):
    try:
        hdrs = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"}
        res  = requests.get(url, headers=hdrs, timeout=15)
        res.raise_for_status()
        soup = BeautifulSoup(res.text, "lxml")
        for tag in soup(["script", "style", "nav", "footer", "head", "noscript", "iframe", "svg", "form", "button"]):
            tag.decompose()
        title       = soup.title.string.strip() if soup.title else "No title"
        meta        = soup.find("meta", attrs={"name": "description"})
        description = meta["content"].strip() if meta and meta.get("content") else ""
        lines       = [l.strip() for l in soup.get_text(separator="\n", strip=True).splitlines() if l.strip()]
        cleaned     = "\n".join(lines)
        if len(cleaned) > 15000:
            cleaned = cleaned[:15000] + "\n\n[... content truncated ...]"
        return {"title": title, "description": description, "url": url, "content": cleaned}
    except requests.exceptions.ConnectionError:
        return {"error": f"Could not connect to {url}."}
    except requests.exceptions.Timeout:
        return {"error": f"Timeout: {url}"}
    except requests.exceptions.HTTPError as e:
        return {"error": f"HTTP error: {e}"}
    except Exception as e:
        return {"error": str(e)}


def is_scrape_request(question):
    url_pattern = re.compile(r'https?://[^\s<>"\']+|www\.[^\s<>"\']+', re.IGNORECASE)
    urls        = url_pattern.findall(question)
    if urls:
        cleaned = []
        for url in urls:
            if url.startswith("www."):
                url = "https://" + url
            cleaned.append(url)
        return cleaned
    return None


def parse_scraped_data_to_rows(ai_response):
    """
    Parse AI response into rows for Excel/CSV export.
    Handles: tab+* Key: Value, * Key: Value, **Key**: Value, numbered lists.
    """

    # ── Strategy 1: Enhanced Text/Markdown Parser ─────────────────────────────
    try:
        rows        = []
        headers_set = []

        # Split into blocks by numbered items: "1. ", "2. " etc.
        blocks = re.split(r'(?m)^\d+\.\s+', ai_response)
        blocks = [b.strip() for b in blocks if len(b.strip()) > 15]

        for block in blocks:
            row = {}

            # Skip section headers like "Real Estate Companies", "Locations"
            # that don't contain property-level key:value pairs
            first_line = block.split('\n')[0].strip()
            # Remove markdown bold markers to get clean title
            clean_title = re.sub(r'\*+', '', first_line).strip().rstrip(':')

            # Extract title if it looks like a property/item name
            if clean_title and len(clean_title) > 3:
                row['Title'] = clean_title

            # Extract ALL key: value pairs from the block
            # Handles:
            #   * Key: Value
            #   \t* Key: Value
            #   \t\t* Key: Value
            #   **Key**: Value
            #   Key: Value
            for line in block.split('\n'):
                line = line.strip()
                # Remove leading asterisks, tabs, dashes
                line = re.sub(r'^[\t\s]*[\*\-]+\s*', '', line).strip()
                # Remove bold markers around key
                line = re.sub(r'\*\*([^*]+)\*\*', r'\1', line)

                # Match "Key: Value" pattern
                match = re.match(r'^([^:\n]{2,40}?)\s*:\s*(.+)$', line)
                if match:
                    key = match.group(1).strip().rstrip(':').strip()
                    val = match.group(2).strip()
                    # Skip noise
                    if (key and val
                            and key.lower() not in ['note', 'notes', 'properties',
                                                     'details', 'found', 'here',
                                                     'real estate companies', 'locations']
                            and len(key) < 50
                            and len(val) < 500):
                        row[key] = val
                        if key not in headers_set:
                            headers_set.append(key)

            # Only keep rows that have meaningful data (more than just a title)
            if len(row) > 1:
                rows.append(row)

        # Only use properties section — skip company/location sub-sections
        # by keeping rows that have at least 3 meaningful fields
        property_rows = [r for r in rows if len(r) >= 3]

        if property_rows and len(property_rows) >= 1:
            # Fill missing keys
            for row in property_rows:
                for h in headers_set:
                    if h not in row:
                        row[h] = ""
            print(f"PARSE SUCCESS (markdown): {len(property_rows)} rows, {len(headers_set)} columns")
            return headers_set, property_rows

    except Exception as e:
        print("MARKDOWN PARSE ERROR:", e)

    # ── Strategy 2: AI JSON Fallback ──────────────────────────────────────────
    try:
        # Clean markdown/code blocks
        cleaned = re.sub(r'^```(?:json)?\s*|\s*```$', '', ai_response.strip(), flags=re.MULTILINE)

        # Try direct JSON parse first
        try:
            direct = json.loads(cleaned)
            if isinstance(direct, list) and direct and isinstance(direct[0], dict):
                headers = list(direct[0].keys())
                print(f"PARSE SUCCESS (direct JSON): {len(direct)} rows")
                return headers, direct
        except json.JSONDecodeError:
            pass

        # Ask AI to convert to JSON
        retry_prompt = (
            "Convert the following scraped data into a VALID JSON array of objects.\n"
            "RULES:\n"
            "1. Return ONLY a JSON array: [{\"key\":\"value\"}, ...]\n"
            "2. NO markdown, NO backticks, NO extra text\n"
            "3. Use consistent keys across all objects\n"
            "4. If a field is missing, use \"\"\n"
            "5. Only include the main records (e.g. properties), not sub-sections\n"
            "6. Return [] if parsing is impossible\n\n"
            "DATA:\n" + cleaned[:3500]
        )

        res = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
            json={
                "model": MODEL,
                "messages": [
                    {"role": "system", "content": "You output ONLY valid JSON arrays. No explanations."},
                    {"role": "user",   "content": retry_prompt}
                ],
                "max_tokens":  4000,
                "temperature": 0.1,
            },
            timeout=30
        )

        raw    = res.json()["choices"][0]["message"]["content"].strip()
        raw    = re.sub(r'^```(?:json)?\s*|\s*```$', '', raw, flags=re.MULTILINE)
        parsed = json.loads(raw)

        if isinstance(parsed, list) and parsed and isinstance(parsed[0], dict):
            headers = list(parsed[0].keys())
            print(f"PARSE SUCCESS (AI retry): {len(parsed)} rows")
            return headers, parsed

    except Exception as e:
        print("AI PARSE ERROR:", e)

    return [], []

# ─────────────────────────────────────────────────────────────────────────────
# Export Helpers
# ─────────────────────────────────────────────────────────────────────────────

def export_to_csv(headers, rows):
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)
    return output.getvalue()


def export_to_excel(headers, rows, title="Scraped Data"):
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = title[:31]
    hf    = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1e3a6e")
    for col_idx, h in enumerate(headers, 1):
        c      = ws.cell(row=1, column=col_idx, value=h)
        c.font = hf
        c.fill = hfill
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=ri, column=ci, value=str(row.get(h, "")))
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 50)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_to_google_sheets(headers, rows, title="Gabriella Scraped Data"):
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = ["https://www.googleapis.com/auth/spreadsheets",
                  "https://www.googleapis.com/auth/drive"]
        creds  = Credentials.from_service_account_file(GSHEET_CREDS, scopes=scopes)
        gc     = gspread.authorize(creds)
        sh     = gc.create(title)
        ws     = sh.get_worksheet(0)
        ws.update([headers] + [[str(row.get(h, "")) for h in headers] for row in rows])
        sh.share(None, perm_type="anyone", role="reader")
        return {"url": sh.url, "title": sh.title}
    except FileNotFoundError:
        return {"error": "google_credentials.json not found."}
    except Exception as e:
        return {"error": str(e)}

# ─────────────────────────────────────────────────────────────────────────────
# Human-like Response Helpers
# ─────────────────────────────────────────────────────────────────────────────

def make_more_human(text):
    """Post-process AI responses to add human-like touches."""
    if not text:  # ✅ Fix: handle empty string
        return text
        
    openers = ["", "Sure! ", "Got it! ", "Okay, ", "Hmm, ", "So, ", "Well, ", "Actually, ", "Oh! "]
    if random.random() < 0.3:
        opener = random.choice(openers[1:])
        # ✅ Fix: check text length before accessing text[0]
        if len(text) > 2 and not text[0].isupper():
            text = opener + text[0].lower() + text[1:]
        elif not text.startswith(opener.strip()):
            text = opener + text
    if random.random() < 0.2:
        emojis = ["👍", "😊", "💡", "✨", "📊", "✅"]
        if text and text[-1] not in ['.', '!', '?', '📊', '✅']:
            text = text.rstrip() + " " + random.choice(emojis)
    replacements = {
        "I can assist you with": "I can help with",
        "Certainly": "Sure",
        "Additionally": "Also",
        "Furthermore": "Plus",
        "However": "But",
        "Therefore": "So",
        "utilize": "use",
        "approximately": "about",
    }
    for formal, casual in replacements.items():
        text = text.replace(formal, casual)
    return text


def detect_conversation_tone(messages):
    if not messages:
        return "neutral"
    recent_user_msgs  = [m["content"].lower() for m in messages[-6:] if m["role"] == "user"]
    casual_indicators = ["lol", "haha", "omg", "thx", "plz", "yeah", "nah", "cool", "awesome"]
    formal_indicators = ["please", "thank you", "could you", "would you", "kindly"]
    casual_score      = sum(1 for msg in recent_user_msgs if any(ci in msg for ci in casual_indicators))
    formal_score      = sum(1 for msg in recent_user_msgs if any(fi in msg for fi in formal_indicators))
    if casual_score > formal_score:
        return "casual"
    elif formal_score > casual_score:
        return "formal"
    return "neutral"

# ─────────────────────────────────────────────────────────────────────────────
# Math Mode Helpers
# ─────────────────────────────────────────────────────────────────────────────

def is_math_query(query):
    math_keywords     = [
        'calculate', 'compute', 'solve', 'equation', 'formula', 'algebra',
        'derivative', 'integral', 'sqrt', 'log', 'ln', 'exp', 'factorial',
        'percentage', 'percent', 'ratio', 'average', 'mean', 'median',
        'sum', 'total', 'difference', 'product', 'quotient', 'remainder',
        'interest', 'emi', 'amortize', 'compound', 'simple interest',
        'x=', 'y=', 'find x', 'solve for', 'what is', 'how much'
    ]
    arithmetic_pattern = r'^[\d\s\+\-\*/\.\(\)\^%]+$'
    equation_pattern   = r'[a-zA-Z]\s*[=\+\-\*/]\s*[\d\s\+\-\*/\.\(\)]+'
    query_lower        = query.lower().strip()
    if re.match(arithmetic_pattern, query_lower.replace(' ', '')):
        return True, "arithmetic"
    if re.search(equation_pattern, query_lower):
        return True, "equation"
    keyword_count = sum(1 for kw in math_keywords if kw in query_lower)
    has_numbers   = bool(re.search(r'\d+', query_lower))
    if keyword_count >= 2 or (keyword_count >= 1 and has_numbers):
        return True, "word_problem"
    return False, None


def safe_arithmetic_eval(expression):
    expr = expression.strip()
    if not all(c in set('0123456789+-*/().^% ') for c in expr):
        return False, "Invalid characters in expression"
    expr = expr.replace('^', '**')
    try:
        tree      = ast.parse(expr, mode='eval')
        operators = {
            ast.Add:  operator.add,
            ast.Sub:  operator.sub,
            ast.Mult: operator.mul,
            ast.Div:  operator.truediv,
            ast.Pow:  operator.pow,
            ast.Mod:  operator.mod,
            ast.UAdd: operator.pos,
            ast.USub: operator.neg,
        }
        def eval_node(node):
            if isinstance(node, ast.Expression):
                return eval_node(node.body)
            # ✅ Fix: Python version compatibility for ast nodes
            elif isinstance(node, (ast.Constant, ast.Num)):
                val = getattr(node, 'value', getattr(node, 'n', 0))
                return Decimal(str(val))
            elif isinstance(node, ast.BinOp):
                left    = eval_node(node.left)
                right   = eval_node(node.right)
                op_type = type(node.op)
                if op_type in operators:
                    if op_type == ast.Div and right == 0:
                        raise ZeroDivisionError("Division by zero")
                    return operators[op_type](left, right)
                raise ValueError(f"Unsupported operator: {op_type}")
            elif isinstance(node, ast.UnaryOp):
                operand = eval_node(node.operand)
                op_type = type(node.op)
                if op_type in operators:
                    return operators[op_type](operand)
                raise ValueError(f"Unsupported unary operator: {op_type}")
            else:
                raise ValueError(f"Unsupported expression type: {type(node)}")
        result = eval_node(tree)
        if result == int(result):
            return True, str(int(result))
        formatted = f"{result:.10f}".rstrip('0').rstrip('.')
        return True, formatted
    except ZeroDivisionError as e:
        return False, f"Error: {str(e)}"
    except (SyntaxError, ValueError, InvalidOperation) as e:
        return False, f"Parse error: {str(e)}"
    except Exception as e:
        return False, f"Calculation error: {str(e)}"


def extract_numbers_and_ops(query):
    patterns = [
        r'([\d.]+)\s*[%\*×x]\s*of\s+([\d.]+)',
        r'([\d.]+)\s*[\+\-\*/]\s*([\d.]+)',
        r'([\d.]+)\s*\^\s*([\d.]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, query, re.IGNORECASE)
        if match:
            groups = match.groups()
            if '%' in pattern or 'of' in query.lower():
                try:
                    pct = Decimal(groups[0]) / Decimal('100')
                    val = Decimal(groups[1])
                    return True, str(pct * val)
                except:
                    pass
            else:
                expr = f"{groups[0]}{match.group(0)[len(groups[0]):len(match.group(0))-len(groups[1])]}{groups[1]}"
                expr = expr.replace('×', '*').replace('x', '*')
                return safe_arithmetic_eval(expr)
    return False, None

# ─────────────────────────────────────────────────────────────────────────────
# Web Search
# ─────────────────────────────────────────────────────────────────────────────

def needs_search(question):
    try:
        res = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
            json={
                "model": MODEL,
                "messages": [
                    {"role": "system", "content": "Does this question need a real-time web search? Reply only YES or NO."},
                    {"role": "user",   "content": question}
                ],
                "max_tokens":  5,
                "temperature": 0.3,
            },
            timeout=(10, 15)
        )
        return res.json()["choices"][0]["message"]["content"].strip().upper().startswith("YES")
    except:
        return False


def web_search(query, max_results=4):
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=max_results))
        if not results:
            return None
        return "\n\n".join(
            f"[{i}] {r.get('title','')}\n{r.get('body','')}\nSource: {r.get('href','')}"
            for i, r in enumerate(results, 1)
        )
    except Exception as e:
        print("SEARCH ERROR:", e)
        return None

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def generate_voice(text):
    tmp      = tempfile.NamedTemporaryFile(delete=False, suffix=".mp3", dir="static", prefix="voice_")
    filename = tmp.name
    tmp.close()
    tts = gTTS(text[:500])
    tts.save(filename)
    return "/" + filename


def read_pdf(file):
    text   = ""
    reader = PyPDF2.PdfReader(file)
    for page in reader.pages:
        text += page.extract_text() or ""
    return text

# ─────────────────────────────────────────────────────────────────────────────
# Complex Math Solver (SymPy)
# ─────────────────────────────────────────────────────────────────────────────

def parse_math_expression(expr_str):
    """Safely parse a math expression string into a SymPy object."""
    transformations = (standard_transformations + (implicit_multiplication_application,))
    try:
        return parse_expr(expr_str, transformations=transformations, evaluate=False)
    except (SympifyError, Exception):
        return None


def solve_complex_math(query):
    """Attempt to solve complex math problems using SymPy."""
    query_lower = query.lower().strip()
    
    # Equation solving
    if '=' in query and re.search(r'[a-zA-Z]', query):
        try:
            if ' = ' in query:
                left, right = query.split(' = ', 1)
            elif '=' in query:
                left, right = query.split('=', 1)
            else:
                return False, "Could not parse equation"
            
            x, y, z, t = symbols('x y z t')
            left_expr = parse_math_expression(left.strip())
            right_expr = parse_math_expression(right.strip())
            
            if left_expr is None or right_expr is None:
                return False, "Could not parse expression"
            
            equation = Eq(left_expr, right_expr)
            vars_to_solve = [v for v in [x, y, z, t] if v in equation.free_symbols]
            if not vars_to_solve:
                return False, "No variables found to solve for"
            
            solution = solve(equation, vars_to_solve[0] if len(vars_to_solve)==1 else vars_to_solve)
            sol_str = ', '.join(str(s) for s in solution) if isinstance(solution, list) else str(solution)
            
            return True, {
                "type": "equation_solution", "query": query, "result": sol_str,
                "steps": [f"Equation: {left} = {right}", f"Solved for: {vars_to_solve[0]}", f"Solution: {sol_str}"],
                "confidence": "high"
            }
        except Exception as e:
            return False, f"Equation solve error: {str(e)}"
    
    # Derivative
    elif any(kw in query_lower for kw in ['derivative of', 'd/dx', 'differentiate']):
        try:
            expr_part = re.sub(r'^(derivative of|d/dx|differentiate)\s*', '', query_lower, flags=re.IGNORECASE).strip()
            expr = parse_math_expression(expr_part)
            if expr is None:
                return False, "Could not parse expression for derivative"
            x = symbols('x')
            result = diff(expr, x)
            return True, {
                "type": "derivative", "query": query, "result": str(result),
                "steps": [f"Expression: {expr_part}", f"Variable: x", f"Derivative: d/dx({expr}) = {result}"],
                "confidence": "high"
            }
        except Exception as e:
            return False, f"Derivative error: {str(e)}"
    
    # Integral
    elif any(kw in query_lower for kw in ['integral of', 'integrate', '∫']):
        try:
            expr_part = re.sub(r'^(integral of|integrate|∫)\s*', '', query_lower, flags=re.IGNORECASE).strip()
            expr = parse_math_expression(expr_part)
            if expr is None:
                return False, "Could not parse expression for integral"
            x = symbols('x')
            result = integrate(expr, x)
            return True, {
                "type": "integral", "query": query, "result": str(result),
                "steps": [f"Expression: {expr_part}", f"Variable: x", f"Integral: ∫({expr})dx = {result}"],
                "confidence": "medium"
            }
        except Exception as e:
            return False, f"Integral error: {str(e)}"
    
    # Simplify/Expand/Factor
    elif any(kw in query_lower for kw in ['simplify', 'expand', 'factor']):
        try:
            expr_part = re.sub(r'^(simplify|expand|factor)\s*', '', query_lower, flags=re.IGNORECASE).strip()
            expr = parse_math_expression(expr_part)
            if expr is None:
                return False, "Could not parse expression"
            if 'simplify' in query_lower:
                result, op_name = simplify(expr), "Simplified"
            elif 'expand' in query_lower:
                result, op_name = expand(expr), "Expanded"
            elif 'factor' in query_lower:
                result, op_name = factor(expr), "Factored"
            else:
                result, op_name = simplify(expr), "Simplified"
            return True, {
                "type": "algebraic_manipulation", "query": query, "result": str(result),
                "steps": [f"Original: {expr_part}", f"Operation: {op_name}", f"Result: {result}"],
                "confidence": "high"
            }
        except Exception as e:
            return False, f"Algebra error: {str(e)}"
    
    # Limit - ✅ Fix: more flexible regex
    elif 'limit' in query_lower and ('as' in query_lower or 'approaches' in query_lower or '->' in query_lower or '→' in query_lower):
        try:
            match = re.match(r'limit of\s+(.+?)\s+as\s+([a-zA-Z])\s*(?:->|approaches|→)\s*([\d\.\-+]+)', query_lower)
            if not match:
                return False, "Could not parse limit expression"
            expr_part, var, val = match.groups()
            expr = parse_math_expression(expr_part.strip())
            if expr is None:
                return False, "Could not parse limit expression"
            sym_var = symbols(var)
            limit_val = float(val) if '.' in val or 'e' in val.lower() else int(val)
            result = limit(expr, sym_var, limit_val)
            return True, {
                "type": "limit", "query": query, "result": str(result),
                "steps": [f"Expression: {expr_part}", f"Limit: {var} → {val}", f"Result: {result}"],
                "confidence": "medium"
            }
        except Exception as e:
            return False, f"Limit error: {str(e)}"
    
    # System of equations
    elif query.count('=') >= 2 and ',' in query:
        try:
            eq_strings = [eq.strip() for eq in re.split(r'[,;]| and ', query) if '=' in eq]
            equations, x, y, z = [], symbols('x y z')
            for eq_str in eq_strings:
                if ' = ' in eq_str:
                    l, r = eq_str.split(' = ', 1)
                else:
                    l, r = eq_str.split('=', 1)
                left, right = parse_math_expression(l.strip()), parse_math_expression(r.strip())
                if left and right:
                    equations.append(Eq(left, right))
            if len(equations) < 2:
                return False, "Need at least 2 valid equations for system"
            solution = solve(equations, [x, y, z])
            return True, {
                "type": "system_solution", "query": query, "result": str(solution),
                "steps": [f"Equations: {', '.join(eq_strings)}", f"Solution: {solution}"],
                "confidence": "high"
            }
        except Exception as e:
            return False, f"System solve error: {str(e)}"
    
    # Generic fallback
    elif re.search(r'[a-zA-Z]', query) and any(op in query for op in ['+', '-', '*', '/', '^', '**']):
        try:
            expr = parse_math_expression(query)
            if expr is None:
                return False, "Could not parse expression"
            x = symbols('x')
            if expr.has(x):
                solution = solve(expr, x)
                return True, {
                    "type": "generic_solve", "query": query, "result": str(solution),
                    "steps": [f"Expression: {query}", f"Solved: {expr} = 0", f"Solution: {solution}"],
                    "confidence": "medium"
                }
        except:
            pass
    
    return False, "Problem type not recognized or too complex for symbolic solver"

# ─────────────────────────────────────────────────────────────────────────────
# PDF Q&A Functions
# ─────────────────────────────────────────────────────────────────────────────

def answer_pdf_questions(pdf_text, user_instruction=""):
    instruction = user_instruction if user_instruction else "Answer all questions found in this document."
    prompt = (
        "You are an expert assistant. The following is the content of a PDF document.\n\n"
        "PDF CONTENT:\n" + pdf_text[:12000] + "\n\n"
        "TASK: " + instruction + "\n\n"
        "Instructions:\n"
        "1. Find ALL questions in the document (numbered, lettered, or plain questions)\n"
        "2. Answer each question clearly and completely\n"
        "3. Return your response as a JSON array with this exact format:\n"
        '[{"question": "full question text", "answer": "complete answer"}]\n\n'
        "If no questions are found, return:\n"
        '[{"question": "Document Summary", "answer": "summary of the document"}]\n\n'
        "Return ONLY valid JSON. No markdown, no backticks, no explanation outside JSON."
    )
    try:
        res = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
            json={
                "model": MODEL, "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 3000, "temperature": 0.7, "top_p": 0.9,
                "presence_penalty": 0.1, "frequency_penalty": 0.1,
            },
            timeout=40
        )
        raw = res.json()["choices"][0]["message"]["content"].strip().replace("```json", "").replace("```", "").strip()
        qa_list = json.loads(raw)
        if isinstance(qa_list, list):
            return qa_list
    except Exception as e:
        print("QA EXTRACTION ERROR:", e)
    return []


def generate_answer_pdf(qa_list, original_filename="document"):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER
    from datetime import datetime

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    title_style    = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1e3a6e'), spaceAfter=6, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#7a8aaa'), spaceAfter=20, alignment=TA_CENTER)
    question_style = ParagraphStyle('Question', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#1a2840'), fontName='Helvetica-Bold', spaceBefore=16, spaceAfter=6, backColor=colors.HexColor('#e8edf5'))
    answer_style   = ParagraphStyle('Answer', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#2d3748'), spaceAfter=8, leftIndent=16, leading=16)
    label_q_style  = ParagraphStyle('LabelQ', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#3b7cf4'), fontName='Helvetica-Bold', spaceAfter=2)
    label_a_style  = ParagraphStyle('LabelA', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#22c55e'), fontName='Helvetica-Bold', spaceAfter=2, leftIndent=16)

    story = []
    story.append(Paragraph("Answered Questions", title_style))
    story.append(Paragraph(f"Source: {original_filename} | Generated by Gabriella AI | {datetime.now().strftime('%B %d, %Y')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1e3a6e'), spaceAfter=20))

    if not qa_list:
        story.append(Paragraph("No questions were found in the uploaded document.", answer_style))
    else:
        for i, qa in enumerate(qa_list, 1):
            story.append(Paragraph(f"QUESTION {i}", label_q_style))
            story.append(Paragraph(qa.get("question", "").strip().replace('\n', '<br/>'), question_style))
            story.append(Paragraph("ANSWER", label_a_style))
            story.append(Paragraph(qa.get("answer", "").strip().replace('\n', '<br/>'), answer_style))
            if i < len(qa_list):
                story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#e2e8f0'), spaceBefore=8, spaceAfter=4))

    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0')))
    story.append(Paragraph(f"Total: {len(qa_list)} question(s) answered | Powered by Gabriella AI", subtitle_style))
    doc.build(story)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────────────────────────────
# Groq Streaming
# ─────────────────────────────────────────────────────────────────────────────

def call_groq_stream(messages):
    with requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
        json={
            "model": MODEL, "messages": messages, "stream": True,
            "temperature": 0.7, "top_p": 0.9, "presence_penalty": 0.1, "frequency_penalty": 0.1,
        },
        stream=True, timeout=(15, 60)
    ) as res:
        for line in res.iter_lines():
            if not line: continue
            decoded = line.decode("utf-8").strip()
            if not decoded.startswith("data: "): continue
            chunk = decoded.replace("data: ", "", 1)
            if chunk == "[DONE]": break
            try:
                cd = json.loads(chunk)
                if "choices" not in cd: continue
                content = cd["choices"][0].get("delta", {}).get("content")
                if content: yield content
            except: continue

# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/solve", methods=["POST"])
def solve():
    """Main chat endpoint with multi-problem handling, progress tracking, and caching."""
    if not API_KEY:
        return jsonify({"answer": "️ API key not configured."}), 500

    request_data = request.get_json()
    question     = request_data.get("question", "").strip()
    math_mode    = request_data.get("math_mode", False)

    if not question:
        return jsonify({"answer": "⚠️ Please enter something"})

    save_message("user", question)
    chat_history = load_history(limit=20)

    # Detect multi-problem messages (separated by newlines)
    lines = [line.strip() for line in question.split('\n') if line.strip()]
    is_multi_math = len(lines) > 1 and all(is_math_query(line)[0] for line in lines)

    def generate():
        full_answer = ""
        try:
            # ─── MULTI-PROBLEM HANDLING ──────────────────────────────────────
            if is_multi_math:
                yield f"📝 I see {len(lines)} math problems. Solving them one by one...\n\n"
                results = []
                for i, prob in enumerate(lines, 1):
                    yield f"⏳ [{i}/{len(lines)}] {prob[:60]}{'...' if len(prob)>60 else ''}\n"

                    # Check cache
                    if prob in MATH_CACHE:
                        res = MATH_CACHE[prob]
                        yield f"✅ (Cached) {res.split('**')[1] if '**' in res else 'Solved'}\n\n"
                        results.append(res)
                        continue

                    # Try local solvers
                    success, result = safe_arithmetic_eval(prob)
                    if not success:
                        success, result = solve_complex_math(prob)

                    if success and isinstance(result, dict):
                        res_text = result.get("result", "Solved")
                        steps = result.get("steps", [])
                        conf = result.get("confidence", "medium")
                        badge = "✅ Exact" if conf == "high" else "⚠️ Approx"
                        formatted = f"🧮 **{res_text}**\n*{badge}*\n" + "\n".join(f"• {s}" for s in steps)
                        yield f"✅ {res_text}\n\n"
                        results.append(formatted)
                        MATH_CACHE[prob] = formatted
                    else:
                        yield "⚠️ Couldn't solve automatically\n\n"
                        results.append(f"⚠️ {prob}")

                full_answer = "\n\n".join(results)
                save_message("assistant", full_answer)
                yield full_answer
                return

            # ─── SINGLE PROBLEM MATH PRE-PROCESSING ──────────────────────────
            is_math, math_type = is_math_query(question)
            math_answer = None

            if math_mode or is_math:
                yield "🧮 Detecting problem type...\n"

                # Check cache first
                if question in MATH_CACHE:
                    math_answer = MATH_CACHE[question]
                    yield "✅ (Cached)\n"
                else:
                    # Priority 1: Simple arithmetic (fast)
                    success, result = safe_arithmetic_eval(question)
                    if success:
                        math_answer = f"🧮 **{result}**\n\n*✅ Exact — calculated with Python's decimal module*"
                        MATH_CACHE[question] = math_answer
                        yield "✅ Calculated exactly\n"
                    else:
                        # Priority 2: Complex symbolic math
                        yield " Attempting symbolic solution...\n"
                        success, result = solve_complex_math(question)
                        if success:
                            res_val = result.get("result", "")
                            steps = result.get("steps", [])
                            conf = result.get("confidence", "medium")
                            steps_text = "\n".join(f"• {s}" for s in steps) if steps else ""
                            badge = "✅ Exact" if conf == "high" else "⚠️ Approx"
                            math_answer = f"🧮 **{res_val}**\n\n*{badge} — solved with SymPy*\n\n{steps_text}"
                            MATH_CACHE[question] = math_answer
                            yield "✅ Solved symbolically\n"
                        else:
                            yield "🤔 Falling back to AI reasoning...\n"

                if math_answer:
                    save_message("assistant", math_answer)
                    yield math_answer
                    return

            # ── AI CONTEXT BUILDING ─────────────────────────────────────────
            messages = list(chat_history)

            # Gabriella's identity
            messages.insert(0, {
                "role": "system",
                "content": (
                    "You're Gabriella, a warm and friendly AI assistant who texts like a real human. "
                    "Use contractions naturally. Keep responses conversational. Vary sentence length. "
                    "Use casual phrases like 'Sure thing!', 'Got it', 'Hmm let me think'. "
                    "Add occasional emojis but don't overdo it. Show enthusiasm. Admit uncertainty naturally. "
                    "For math/technical stuff: explain simply. Be professional but approachable. "
                    "Your name is Gabriella. Always introduce yourself as Gabriella when asked. "
                    "AVOID: overly formal language, repeating phrases, long paragraphs, textbook tone, 'As an AI...'"
                )
            })

            # Bank sheet context
            if bank_sheet_context["summary"]:
                messages.insert(1, {
                    "role": "system",
                    "content": "Bank sheet summary:\n\n" + bank_sheet_context["summary"] + "\n\nBe specific with numbers."
                })

            # Tone adaptation
            tone = detect_conversation_tone(messages)
            if tone == "casual":
                tone_instruction = "User is casual — match their energy! Use emojis, slang, relaxed language."
            elif tone == "formal":
                tone_instruction = "User is formal — keep it professional and polite."
            else:
                tone_instruction = "Be friendly and conversational, but professional."
            messages.insert(2, {"role": "system", "content": tone_instruction})

            # ─── SCRAPING / WEB SEARCH ───────────────────────────────────────
            scraped_urls = is_scrape_request(question)
            all_scraped_data = []

            if scraped_urls:
                yield f"🌐 Scraping {len(scraped_urls)} page(s)...\n"
                all_scraped_content = ""
                for i, url in enumerate(scraped_urls, 1):
                    yield f"⏳ Page {i}/{len(scraped_urls)}: {url}\n"
                    scraped = scrape_url(url)
                    if "error" in scraped:
                        yield f"️ Failed: {scraped['error']}\n"
                        continue
                    all_scraped_data.append(scraped)
                    all_scraped_content += f"\n\n--- PAGE {i}: {scraped['title']} ---\n" + scraped['content']

                if not all_scraped_data:
                    yield "❌ All pages failed to scrape."
                    return

                yield "📊 Extracting structured data...\n"
                messages.append({
                    "role": "user",
                    "content": (
                        f"I scraped {len(all_scraped_data)} webpage(s):\n" + all_scraped_content[:15000] +
                        f"\n\nUser request: {question}\n\n"
                        f"Extract and organize ALL relevant data from ALL pages combined. "
                        f"Do not duplicate records. Look for: names, addresses, phones, emails, websites, hours, "
                        f"prices, product codes, locations, social links. Present ALL records in clean structured format. "
                        f"If a field is not found, write 'Not found'. Only extract data that actually exists."
                    )
                })

            elif needs_search(question):
                yield "🔍 Searching the web...\n"
                results = web_search(question)
                if results:
                    messages.append({"role": "user", "content": f"Web results:\n\n{results}\n\nAnswer: {question}"})
                else:
                    yield "⚠️ Web search unavailable.\n"

            # ── STREAM AI RESPONSE (FIXED: no post-processing on chunks) ────
            yield "💭 Generating response...\n"
            for content in call_groq_stream(messages):
                full_answer += content
                yield content  # ✅ Stream raw AI output (personality handled via system prompt)

            # ── POST-STREAM EXPORT HANDLING ─────────────────────────────────
            if scraped_urls and full_answer:
                hdrs, rows = parse_scraped_data_to_rows(full_answer)
                combined_title = all_scraped_data[0].get("title", "Scraped Data") if all_scraped_data else "Scraped Data"
                scraped_data_context.update({"headers": hdrs, "rows": rows, "title": combined_title, "url": str(scraped_urls)})
                if rows:
                    yield f"\n\n✅ Data ready to export — {len(rows)} records from {len(all_scraped_data)} page(s). Use the buttons below."
                else:
                    fallback_rows = [{"Data": line.strip()} for line in full_answer.split('\n') if line.strip()]
                    fallback_title = all_scraped_data[0].get("title", "Scraped Data") if all_scraped_data else "Scraped Data"
                    scraped_data_context.update({"headers": ["Data"], "rows": fallback_rows, "title": fallback_title, "url": str(scraped_urls)})
                    yield "\n\n✅ Data ready to export — use the buttons below."

            if full_answer:
                save_message("assistant", full_answer)

        except Exception as e:
            print("STREAM ERROR:", e)
            yield f"\n\n⚠️ Error: {str(e)}"

    return Response(generate(), mimetype="text/plain")


@app.route("/math/export_excel", methods=["POST"])
def math_export_excel():
    """✅ FIXED: Generate Excel file for math verification."""
    try:
        request_data = request.get_json()
        query   = request_data.get("query", "Calculation")
        result  = request_data.get("result", "")
        steps   = request_data.get("steps", [])
        result_type = request_data.get("result_type", "")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Math Verification"
        header_fill = PatternFill("solid", fgColor="1e3a6e")
        header_font = Font(bold=True, color="FFFFFF")

        ws['A1'], ws['B1'] = "Query", query
        ws['A1'].font, ws['A1'].fill = header_font, header_fill
        ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 15, 50

        ws['A3'], ws['B3'] = "Result", result
        ws['A3'].font = Font(bold=True)

        # ✅ FIXED: Properly indented inside try block
        if steps:
            ws['A5'] = "Steps"
            ws['A5'].font = Font(bold=True)
            for i, step in enumerate(steps, 6):
                ws[f'A{i}'], ws[f'B{i}'] = f"{i-5}.", step

        # ✅ FIXED: Define last_row before using it
        last_row = 7 if not steps else len(steps) + 5
        
        # SymPy-specific verification hints
        if result_type in ["equation_solution", "system_solution"]:
            ws[f'A{last_row+2}'] = "🔍 Verify:"
            ws[f'B{last_row+2}'] = "Plug solutions back into original equation(s)"
            ws[f'B{last_row+2}'].font = Font(italic=True, color="666666")
        elif result_type == "derivative":
            ws[f'A{last_row+2}'] = "🔍 Verify:"
            ws[f'B{last_row+2}'] = "Use numerical approximation or graph to confirm slope"
            ws[f'B{last_row+2}'].font = Font(italic=True, color="666666")
        elif result_type == "integral":
            ws[f'A{last_row+2}'] = "🔍 Verify:"
            ws[f'B{last_row+2}'] = "Differentiate the result — should return original function"
            ws[f'B{last_row+2}'].font = Font(italic=True, color="666666")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"math_verification_{hash(query) % 10000}.xlsx")
    except Exception as e:
        print("MATH EXPORT ERROR:", e)
        return jsonify({"error": "Could not generate verification file"}), 500


@app.route("/detect_errors", methods=["POST"])
def detect_errors_route():
    headers = bank_sheet_context.get("headers", [])
    records = bank_sheet_context.get("records", [])
    summary = bank_sheet_context.get("summary", "")
    preview = bank_sheet_context.get("raw_preview", "")
    if not records:
        return jsonify({"error": "No bank sheet loaded. Please upload one first."}), 400
    fixes = detect_errors(headers, records, summary, preview)
    correction_state["pending_fixes"], correction_state["accepted_fixes"], correction_state["done"] = fixes, [], False
    return jsonify({"total": len(fixes), "fixes": fixes, "message": f"Found {len(fixes)} potential issue(s). Review each one below."})


@app.route("/review_fix", methods=["POST"])
def review_fix():
    data, fix_id, decision = request.get_json(), request.get_json().get("fix_id"), request.get_json().get("decision")
    fix = next((f for f in correction_state["pending_fixes"] if f["id"] == fix_id), None)
    if not fix:
        return jsonify({"error": "Fix not found."}), 404
    if decision == "accept":
        correction_state["accepted_fixes"].append(fix)
    remaining = [f for f in correction_state["pending_fixes"] if f["id"] != fix_id and f["id"] not in {a["id"] for a in correction_state["accepted_fixes"]}]
    return jsonify({"status": decision, "accepted_so_far": len(correction_state["accepted_fixes"]), "remaining": len(remaining)})


@app.route("/download/corrected", methods=["GET"])
def download_corrected():
    headers, records, accepted = bank_sheet_context.get("headers", []), bank_sheet_context.get("records", []), correction_state.get("accepted_fixes", [])
    if not records:
        return jsonify({"error": "No bank sheet loaded."}), 400
    buf = build_corrected_excel(headers, records, accepted)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="corrected_bank_sheet.xlsx")


@app.route("/export/csv")
def export_csv():
    headers, rows = scraped_data_context.get("headers", []), scraped_data_context.get("rows", [])
    if not rows:
        return jsonify({"error": "No scraped data. Scrape a URL first."}), 400
    return Response(export_to_csv(headers, rows), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=scraped_data.csv"})


@app.route("/export/excel")
def export_excel_route():
    headers, rows, title = scraped_data_context.get("headers", []), scraped_data_context.get("rows", []), scraped_data_context.get("title", "Scraped Data")
    if not rows:
        return jsonify({"error": "No scraped data. Scrape a URL first."}), 400
    return send_file(export_to_excel(headers, rows, title), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="scraped_data.xlsx")


@app.route("/export/gsheet")
def export_gsheet():
    headers, rows, title = scraped_data_context.get("headers", []), scraped_data_context.get("rows", []), scraped_data_context.get("title", "Gabriella Scraped Data")
    if not rows:
        return jsonify({"error": "No scraped data."}), 400
    result = export_to_google_sheets(headers, rows, title)
    return jsonify(result) if "error" not in result else (jsonify(result), 500)


@app.route("/upload", methods=["POST"])
def upload():
    if not API_KEY:
        return jsonify({"answer": "⚠️ API key not configured."}), 500
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"answer": "⚠️ No file received."}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"answer": f"❌ File type '{ext}' not allowed."}), 400

    user_message = request.form.get("message", "").strip()

    if ext in (".xlsx", ".xls"):
        try:
            headers, records, stats, preview = parse_excel(file)
        except Exception:
            return jsonify({"answer": "❌ Could not read Excel file."}), 400
        if stats is None:
            return jsonify({"answer": "❌ Empty Excel file."}), 400

        stats_text = format_stats_for_ai(stats)
        bank_sheet_context.update({"summary": stats_text, "raw_preview": preview, "headers": headers, "records": records})
        user_instruction = f"\n\nAdditional instruction from user: {user_message}" if user_message else ""
        prompt = (
            "You are a professional bank data analyst. Statistical summary:\n\n" + stats_text +
            "\n\nPreview:\n\n" + (preview[:2000] if preview else "") +
            "\n\nProvide a comprehensive analysis covering:\n"
            "1. Portfolio overview\n2. Loan distribution\n3. Repayment status\n"
            "4. Defaulter analysis\n5. Red flags\n6. Recommendations\n\nBe specific with numbers." + user_instruction
        )
        try:
            res = requests.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
                json={"model": MODEL, "messages": [{"role": "user", "content": prompt}],
                      "temperature": 0.7, "top_p": 0.9, "presence_penalty": 0.1, "frequency_penalty": 0.1},
                timeout=30
            )
            answer = res.json()["choices"][0]["message"]["content"] if "choices" in res.json() else "❌ API Error"
        except Exception:
            answer = "❌ Failed to analyze."

        save_message("user", f"[Uploaded: {file.filename}]")
        save_message("assistant", answer)
        voice_url = generate_voice(f"Sheet analyzed. {stats['total_records']} records found.")
        return jsonify({"answer": answer, "voice": voice_url, "has_sheet": True})

    try:
        content = read_pdf(file) if ext == ".pdf" else file.read().decode("utf-8", errors="replace")
    except:
        return jsonify({"answer": "❌ Could not read file."}), 400

    try:
        instruction = user_message if user_message else "Summarize this"
        res = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
            json={"model": MODEL, "messages": [{"role": "user", "content": f"{instruction}:\n" + content[:3000]}]},
            timeout=10
        )
        answer = res.json()["choices"][0]["message"]["content"] if "choices" in res.json() else "❌ API Error"
    except:
        answer = "❌ Failed to process file."

    return jsonify({"answer": answer, "voice": generate_voice(answer)})


@app.route("/clear", methods=["POST"])
def clear():
    try:
        clear_history()
        bank_sheet_context.update({"summary": None, "raw_preview": None, "headers": [], "records": []})
        scraped_data_context.update({"rows": [], "headers": [], "title": "", "url": ""})
        correction_state.update({"pending_fixes": [], "accepted_fixes": [], "done": False})
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/answer_pdf", methods=["POST"])
def answer_pdf():
    if not API_KEY:
        return jsonify({"error": "API key not configured."}), 500
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "No file received."}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext != ".pdf":
        return jsonify({"error": "Only PDF files are supported for this feature."}), 400
    user_instruction = request.form.get("message", "").strip()
    try:
        pdf_text = read_pdf(file)
    except Exception as e:
        print("PDF READ ERROR:", e)
        return jsonify({"error": "Could not read the PDF file."}), 400
    if not pdf_text.strip():
        return jsonify({"error": "The PDF appears to be empty or image-based. Only text-based PDFs are supported."}), 400
    qa_list = answer_pdf_questions(pdf_text, user_instruction)
    if not qa_list:
        return jsonify({"error": "Could not extract questions from the PDF."}), 400
    try:
        original_name = os.path.splitext(file.filename)[0]
        pdf_buf = generate_answer_pdf(qa_list, original_name)
    except Exception as e:
        print("PDF GENERATION ERROR:", e)
        return jsonify({"error": "Could not generate the answer PDF."}), 500
    return send_file(pdf_buf, mimetype="application/pdf", as_attachment=True, download_name=f"{original_name}_answered.pdf")


@app.route("/")
def home():
    return render_template("index.html")

# ─────────────────────────────────────────────────────────────────────────────
# Plot / Chart Generation
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/plot/auto", methods=["POST"])
def plot_auto():
    """Use AI to parse a natural language plot request into structured chart data."""
    if not API_KEY:
        return jsonify({"error": "API key not configured."}), 500

    data  = request.get_json()
    query = data.get("query", "").strip()
    if not query:
        return jsonify({"error": "No query provided."}), 400

    prompt = (
        f'Parse this chart/plot request and return ONLY valid JSON:\n'
        f'Query: "{query}"\n\n'
        'Rules:\n'
        '1. chart_type: "line" for mathematical functions (y=f(x)), "bar" for comparisons, "pie" for proportions/percentages\n'
        '2. For LINE charts: expression = valid Python/SymPy math expression using "x" as variable (e.g. "x**2 + 2*x - 1", "sin(x)", "exp(-x)")\n'
        '3. For BAR/PIE charts: items = array of {"label": "...", "value": number}\n'
        '4. x_min, x_max: sensible range for line charts (default -10 to 10, use 0 to 6.28 for trig)\n'
        '5. title: short descriptive title\n\n'
        'Return this exact JSON structure:\n'
        '{"chart_type": "line"|"bar"|"pie", "expression": "...", "items": [], '
        '"x_min": -10, "x_max": 10, "title": "..."}\n\n'
        'Return ONLY valid JSON. No markdown, no backticks, no explanation.'
    )

    try:
        res = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
            json={
                "model": MODEL,
                "messages": [
                    {"role": "system", "content": "You output ONLY valid JSON. No explanations, no markdown."},
                    {"role": "user",   "content": prompt}
                ],
                "max_tokens": 500, "temperature": 0.1,
            },
            timeout=20
        )
        raw    = res.json()["choices"][0]["message"]["content"].strip()
        raw    = raw.replace("```json", "").replace("```", "").strip()
        parsed = json.loads(raw)
    except Exception as e:
        return jsonify({"error": f"Could not parse plot request: {str(e)}"}), 400

    # Hand off to the actual plot generator
    return _generate_chart(parsed)


@app.route("/plot", methods=["POST"])
def plot():
    """Generate chart data directly from structured input."""
    data = request.get_json()
    return _generate_chart(data)


def _generate_chart(data):
    """Core chart data generator — shared by /plot and /plot/auto."""
    chart_type = data.get("chart_type", "line")
    title      = data.get("title", "Chart")
    PALETTE    = ["#3b7cf4","#22c55e","#f97316","#a855f7","#ef4444","#eab308","#06b6d4","#ec4899","#14b8a6","#f43f5e"]

    # ── LINE CHART ────────────────────────────────────────────────────────────
    if chart_type == "line":
        expr_str = data.get("expression", "x")
        x_min    = float(data.get("x_min", -10))
        x_max    = float(data.get("x_max",  10))
        try:
            from sympy import symbols as sym_symbols
            from sympy.parsing.sympy_parser import parse_expr, standard_transformations, implicit_multiplication_application
            x        = sym_symbols('x')
            trans    = standard_transformations + (implicit_multiplication_application,)
            expr     = parse_expr(expr_str, transformations=trans)
            f        = lambdify(x, expr, modules=['numpy'])
            x_vals   = np.linspace(x_min, x_max, 400)
            y_vals   = f(x_vals)

            # Handle scalar output
            if np.isscalar(y_vals):
                y_vals = np.full_like(x_vals, float(y_vals))
            y_vals = np.array(y_vals, dtype=float)

            # Clip extreme values for readability
            y_median = np.nanmedian(y_vals[np.isfinite(y_vals)]) if np.any(np.isfinite(y_vals)) else 0
            y_vals   = np.where(np.abs(y_vals - y_median) > 1e6, np.nan, y_vals)

            labels   = [round(float(v), 3) for v in x_vals]
            y_clean  = [None if not np.isfinite(v) else round(float(v), 6) for v in y_vals]

            return jsonify({
                "chart_type": "line",
                "title": title or f"y = {expr_str}",
                "labels": labels,
                "datasets": [{
                    "label": f"y = {expr_str}",
                    "data": y_clean,
                    "borderColor": "#3b7cf4",
                    "backgroundColor": "rgba(59,124,244,0.08)",
                    "tension": 0.3,
                    "pointRadius": 0,
                    "borderWidth": 2,
                    "fill": True,
                    "spanGaps": True
                }],
                "x_label": "x",
                "y_label": "y"
            })
        except Exception as e:
            return jsonify({"error": f"Could not evaluate expression '{expr_str}': {str(e)}"}), 400

    # ── BAR CHART ─────────────────────────────────────────────────────────────
    elif chart_type == "bar":
        items  = data.get("items", [])
        # Fallback: parse "A:30, B:50" format
        if not items and data.get("expression"):
            for part in data["expression"].split(","):
                if ":" in part:
                    lbl, val = part.split(":", 1)
                    try: items.append({"label": lbl.strip(), "value": float(val.strip())})
                    except: pass
        if not items:
            return jsonify({"error": "No data items provided for bar chart."}), 400

        labels = [it["label"] for it in items]
        values = [it["value"] for it in items]
        colors = [PALETTE[i % len(PALETTE)] for i in range(len(items))]

        return jsonify({
            "chart_type": "bar",
            "title": title,
            "labels": labels,
            "datasets": [{
                "label": title,
                "data": values,
                "backgroundColor": colors,
                "borderColor": [c + "cc" for c in colors],
                "borderWidth": 1,
                "borderRadius": 6
            }],
            "x_label": "", "y_label": "Value"
        })

    # ── PIE CHART ─────────────────────────────────────────────────────────────
    elif chart_type == "pie":
        items  = data.get("items", [])
        if not items and data.get("expression"):
            for part in data["expression"].split(","):
                if ":" in part:
                    lbl, val = part.split(":", 1)
                    try: items.append({"label": lbl.strip(), "value": float(val.strip())})
                    except: pass
        if not items:
            return jsonify({"error": "No data items provided for pie chart."}), 400

        labels = [it["label"] for it in items]
        values = [it["value"] for it in items]
        colors = [PALETTE[i % len(PALETTE)] for i in range(len(items))]

        return jsonify({
            "chart_type": "pie",
            "title": title,
            "labels": labels,
            "datasets": [{
                "data": values,
                "backgroundColor": colors,
                "hoverOffset": 8
            }]
        })

    return jsonify({"error": f"Unknown chart type: {chart_type}"}), 400


if __name__ == "__main__":
    app.run(debug=os.environ.get("FLASK_DEBUG", "false").lower() == "true")